from .forms import ProfileUpdateForm, CustomPasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.http import JsonResponse, FileResponse
from django.contrib import messages
from django.conf import settings
from .models import Movie, Session, Ticket, Hall, FavoriteMovie, Review, BonusTransaction
from datetime import datetime, timedelta, date
import io
import os
import qrcode
from reportlab.lib.pagesizes import A5
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors
from collections import defaultdict
import json
from django.contrib.admin.views.decorators import staff_member_required
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from .forms import CustomUserCreationForm
from .forms import ReviewForm
from django.contrib.auth.decorators import login_required


def export_analytics_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="analytics.pdf"'

    p = canvas.Canvas(response)
    p.drawString(100, 800, "Звіт аналітики продажів")
    p.drawString(100, 780, "Це приклад PDF-файлу з Django.")
    p.showPage()
    p.save()
    return response



def export_analytics_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Продажі"

    ws.append(["Дата", "Кількість квитків"])
    ws['A1'].font = ws['B1'].font = Font(bold=True)

    from .models import Ticket
    from django.db.models import Count
    sales = Ticket.objects.values('session__date').annotate(total=Count('id')).order_by('session__date')
    for s in sales:
        ws.append([s['session__date'], s['total']])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="analytics.xlsx"'
    wb.save(response)
    return response

def movie_list(request):
    movies = Movie.objects.all()

    # Отримаємо всі жанри в один список
    raw_genres = Movie.objects.values_list('genre', flat=True)
    split_genres = list(
        set(
            genre.strip()
            for entry in raw_genres if entry
            for genre in entry.split(',')
        )
    )
    split_genres.sort()

    favorite_ids = []
    if request.user.is_authenticated:
        favorite_ids = FavoriteMovie.objects.filter(user=request.user).values_list('movie_id', flat=True)

    return render(request, 'cinema/movie_list.html', {
        'movies': movies,
        'genres': split_genres,
        'favorite_ids': favorite_ids,
    })


def movie_sessions(request, movie_id):
    movie = get_object_or_404(Movie, id=movie_id)
    today = datetime.today().date()
    dates = [today + timedelta(days=i) for i in range(5)]  # найближчі 5 днів

    selected_date = request.GET.get('date')
    if selected_date:
        selected_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
    else:
        selected_date = today

    sessions = Session.objects.filter(movie=movie, date=selected_date).order_by('start_time')

    sessions_by_hall = {}
    for session in sessions:
        hall_name = session.hall.name
        if hall_name not in sessions_by_hall:
            sessions_by_hall[hall_name] = []
        sessions_by_hall[hall_name].append(session)

    return render(request, 'cinema/movie_sessions.html', {
        'movie': movie,
        'dates': dates,
        'selected_date': selected_date,
        'sessions_by_hall': sessions_by_hall
    })

def movie_detail(request, movie_id):
    movie = get_object_or_404(Movie, id=movie_id)
    today = date.today()
    if request.user.is_authenticated:
        all_reviews = Review.objects.filter(movie=movie).exclude(user=request.user)
    else:
        all_reviews = Review.objects.filter(movie=movie)

    # Доступні сеанси
    sessions = Session.objects.filter(movie=movie, date__gte=today).order_by("date", "start_time")
    unique_dates = sorted(set(session.date for session in sessions))

    # Групування по датах → по залах
    sessions_by_date = defaultdict(lambda: defaultdict(list))
    for session in sessions:
        sessions_by_date[session.date][session.hall.id].append(session)

    # Всі зали
    halls = Hall.objects.all()


    # ВІДГУКИ
    review = None
    review_form = None
    can_review = False

    if request.user.is_authenticated:
        review = Review.objects.filter(user=request.user, movie=movie).first()
        watched = Ticket.objects.filter(
            user=request.user,
            session__movie=movie,
            session__date__lt=today,
            status="Куплений"
        ).exists()
        can_review = watched and review is None

        if request.method == 'POST':
            review_form = ReviewForm(request.POST)
            if review_form.is_valid():
                new_review = review_form.save(commit=False)
                new_review.user = request.user
                new_review.movie = movie
                new_review.save()
                messages.success(request, "✅ Відгук додано!")
                return redirect('movie_detail', movie_id=movie.id)
        else:
            review_form = ReviewForm()

    return render(request, "cinema/movie_detail.html", {
        "movie": movie,
        "sessions_by_date": sessions_by_date,
        "unique_dates": unique_dates,
        "halls": halls,
        "review": review,
        "can_review": can_review,
        "review_form": review_form,
        "all_reviews": all_reviews,
    })

def autocomplete_movies(request):
    term = request.GET.get('term', '').lower()
    if term:
        movies = Movie.objects.filter(title__icontains=term)
        results = []
        for movie in movies:
            results.append({
                'title': movie.title,
                'url': f'/movie/{movie.id}/',  # або твій шлях до сторінки фільму
                'poster': movie.poster.url if movie.poster else '/static/images/no-poster.jpg'
            })
        return JsonResponse(results, safe=False)
    return JsonResponse([], safe=False)
def booking(request, session_id):
    session = get_object_or_404(Session, id=session_id)

    # Отримання заброньованих або куплених місць
    booked_tickets = Ticket.objects.filter(session=session, status__in=["Заброньований", "Куплений"]).values_list('seat_number', flat=True)
    booked_seats = set(booked_tickets)

    rows = list(range(1, 11))
    seats = list(range(1, 11))

    seat_map = []
    for row in rows:
        row_seats = []
        for seat in seats:
            seat_number = (row - 1) * 10 + seat
            row_seats.append({
                "number": seat_number,
                "booked": seat_number in booked_seats
            })
        seat_map.append({"row": row, "seats": row_seats})

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            selected_seats = data.get("seats", [])

            if not selected_seats:
                return JsonResponse({"success": False, "error": "Не вибрано жодного місця!"})

            # перевірка: чи всі місця ще вільні
            for seat in selected_seats:
                if seat in booked_seats:
                    return JsonResponse({"success": False, "error": f"Місце {seat} вже зайняте!"})

            # зберігаємо у сесії для подальшої оплати
            request.session['selected_seats'] = selected_seats
            request.session['session_id'] = session.id

            return JsonResponse({"success": True, "redirect_url": reverse('checkout', args=[session.id])})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return render(request, 'cinema/booking.html', {
        'session': session,
        'seat_map': seat_map,
        'hide_footer': True
    })



def checkout(request, session_id):
    session = get_object_or_404(Session, id=session_id)
    selected_ids = request.session.get("selected_seats", [])
    bonus_discount_applied = request.session.get("bonus_discount", False)

    selected_tickets = []
    total_price = 0

    for seat in selected_ids:
        seat = int(seat)
        row = (seat - 1) // 10 + 1
        seat_number = (seat - 1) % 10 + 1
        price = 250 if seat > 90 else 150
        type_name = "VIP" if seat > 90 else "Стандарт"

        selected_tickets.append({
            "seat_id": seat,
            "row": row,
            "seat": seat_number,
            "type": type_name,
            "price": price
        })

        total_price += price

    # Збираємо бонуси
    bonus_points = 0
    if request.user.is_authenticated:
        bonus_points = sum(BonusTransaction.objects.filter(user=request.user).values_list('amount', flat=True))

    max_bonus = min(total_price, bonus_points)

    request.session["booking_start"] = datetime.now().isoformat()

    return render(request, 'cinema/checkout.html', {
        'session': session,
        'selected_tickets': selected_tickets,
        'total_price': total_price,
        'bonus_points': bonus_points,
        'max_bonus': max_bonus,
    })




def booking_success(request):
    ticket_ids = request.session.get("confirmed_tickets", [])
    booked_tickets = Ticket.objects.filter(id__in=ticket_ids)

    return render(request, 'cinema/booking_success.html', {
        'tickets': booked_tickets
    })



def confirm_checkout(request, session_id):
    session = get_object_or_404(Session, id=session_id)
    selected_ids = request.session.get("selected_seats", [])

    if request.method == "POST":
        try:
            data = json.loads(request.body)
            bonus_amount = int(data.get("bonus_amount", 0))

            tickets = []
            total_price = 0

            for seat in selected_ids:
                seat = int(seat)
                price = 250 if seat > 90 else 150
                total_price += price

                ticket = Ticket(
                    session=session,
                    seat_number=seat,
                    status="Куплений",
                    price=price
                )
                if request.user.is_authenticated:
                    ticket.user = request.user
                ticket.save()
                tickets.append(ticket)

            used_bonus = 0
            if request.user.is_authenticated and bonus_amount > 0:
                current_bonus = sum(BonusTransaction.objects.filter(user=request.user).values_list('amount', flat=True))
                allowed_bonus = min(current_bonus, total_price)
                used_bonus = min(bonus_amount, allowed_bonus)

                if used_bonus > 0:
                    new_balance = current_bonus - used_bonus
                    BonusTransaction.objects.create(
                        user=request.user,
                        description=f"Використано {used_bonus} бонусів при купівлі",
                        amount=-used_bonus,
                        balance_after=new_balance
                    )

            # Сума, що справді оплачена грошима
            paid_amount = total_price - used_bonus

            if request.user.is_authenticated:
                bonus_to_add = int(paid_amount * 0.1)
                if bonus_to_add > 0:
                    add_bonus_points(
                        request.user,
                        bonus_to_add,
                        "Бонуси за покупку квитків"
                    )

            request.session["confirmed_tickets"] = [t.id for t in tickets]
            request.session.pop("selected_seats", None)

            return JsonResponse({"success": True, "redirect_url": reverse("booking_success")})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid method"})





def download_ticket_pdf(request, ticket_id):
    ticket = Ticket.objects.get(id=ticket_id)

    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A5)
    width, height = A5

    # Реєстрація шрифту
    font_path = os.path.join(settings.BASE_DIR, "static/fonts/DejaVuSans.ttf")
    pdfmetrics.registerFont(TTFont("DejaVu", font_path))
    c.setFont("DejaVu", 12)

    # Фон
    c.setFillColor(colors.lightgrey)
    c.rect(0, 0, width, height, fill=1)
    c.setFillColor(colors.black)

    # Рамка
    c.setLineWidth(2)
    c.setStrokeColor(colors.darkblue)
    c.rect(10 * mm, 10 * mm, width - 20 * mm, height - 20 * mm)

    # Логотип
    logo_path = os.path.join(settings.BASE_DIR, "static/images/logo.png")
    if os.path.exists(logo_path):
        logo = ImageReader(logo_path)
        c.drawImage(logo, 15 * mm, height - 30 * mm, width=40 * mm, preserveAspectRatio=True)

    # Заголовок
    c.setFont("DejaVu", 16)
    c.drawCentredString(width / 2, height - 20 * mm, "Квиток до кінотеатру")

    # Інформація
    c.setFont("DejaVu", 12)
    text_lines = [
        f"Фільм: {ticket.session.movie.title}",
        f"Дата: {ticket.session.date.strftime('%d.%m.%Y')}",
        f"Час: {ticket.session.start_time.strftime('%H:%M')} – {ticket.session.end_time.strftime('%H:%M')}",
        f"Місце: Ряд {(ticket.seat_number - 1)//10 + 1}, Місце {(ticket.seat_number - 1)%10 + 1}",
        f"Тип: {'VIP' if ticket.seat_number > 90 else 'Стандарт'}",
        f"Ціна: {ticket.price} грн",
        f"Номер квитка: {ticket.id}"
    ]

    y_position = height - 50 * mm
    for line in text_lines:
        c.drawString(20 * mm, y_position, line)
        y_position -= 10 * mm

    # QR-код
    qr_data = f"http://127.0.0.1:8000/ticket/{ticket.id}/pdf/"
    qr_img = qrcode.make(qr_data)
    qr_buffer = io.BytesIO()
    qr_img.save(qr_buffer, format='PNG')
    qr_buffer.seek(0)
    qr_image = ImageReader(qr_buffer)
    c.drawImage(qr_image, width - 50 * mm, 20 * mm, width=30 * mm, height=30 * mm, preserveAspectRatio=True)

    c.showPage()
    c.save()
    buffer.seek(0)

    return FileResponse(buffer, filename=f"ticket_{ticket.id}.pdf", content_type='application/pdf')


@staff_member_required
def admin_dashboard(request):
    return render(request, 'cinema/dashboard.html')


def analytics_view(request):
    return render(request, "admin/analytics_dashboard.html")

from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib.auth.forms import AuthenticationForm, UserCreationForm
from .models import Ticket
from django.contrib.auth import login as auth_login
from datetime import date

def custom_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            login(request, form.get_user())
            return redirect('user_dashboard')
    else:
        form = AuthenticationForm()
    return render(request, 'cinema/login.html', {'form': form})

def custom_logout(request):
    logout(request)
    return redirect('login')

def register(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            auth_login(request, user)
            return redirect('user_dashboard')
    else:
        form = CustomUserCreationForm()
    return render(request, 'cinema/register.html', {'form': form})

@login_required(login_url='login')
def user_dashboard(request):
    tickets = Ticket.objects.filter(user=request.user).select_related('session__movie', 'session__hall')
    now = datetime.now()

    for ticket in tickets:
        seat_num = ticket.seat_number
        ticket.row = (seat_num - 1) // 10 + 1
        ticket.seat = (seat_num - 1) % 10 + 1

        session_datetime = datetime.combine(ticket.session.date, ticket.session.start_time)
        if ticket.status in ["Куплений", "Активний"] and session_datetime < now:
            ticket.visited_label = "Відвідано"
        else:
            ticket.visited_label = "Очікується"

    user = request.user
    allowed_groups = {'Директор', 'Адміністратор', 'Касир'}
    is_admin_user = user.is_superuser or user.groups.filter(name__in=allowed_groups).exists()
    favorite_movies = Movie.objects.filter(favoritemovie__user=request.user)
    reviews = Review.objects.filter(user=request.user).select_related('movie')
    bonus_points = sum(BonusTransaction.objects.filter(user=request.user).values_list('amount', flat=True))
    bonus_history = BonusTransaction.objects.filter(user=request.user)
    user_level = get_user_status(bonus_points)
    user_status_class = get_user_status_class(bonus_points)
    progress_percent = min(100, int((bonus_points / 50) * 100))
    bonus_progress = min(int((bonus_points / 1000) * 100), 100)

    return render(request, "cinema/account.html", {
        "tickets": tickets,
        "bonus_points": bonus_points,
        'bonus_history': bonus_history,
        'is_admin_user': is_admin_user,
        "favorite_movies": favorite_movies,
        "reviews": reviews,
        "user_level": user_level,
        "user_status_class": user_status_class,
        "progress_percent": progress_percent,
        "bonus_progress": bonus_progress,
    })

@login_required
def toggle_favorite(request, movie_id):
    movie = Movie.objects.get(id=movie_id)
    fav, created = FavoriteMovie.objects.get_or_create(user=request.user, movie=movie)
    if not created:
        fav.delete()
        return JsonResponse({"status": "removed"})
    return JsonResponse({"status": "added"})

@login_required
def update_review(request, movie_id):
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        review = get_object_or_404(Review, user=request.user, movie_id=movie_id)
        form = ReviewForm(request.POST, instance=review)
        if form.is_valid():
            updated_review = form.save()
            return JsonResponse({
                "success": True,
                "text": updated_review.text,
                "rating": updated_review.rating
            })
        return JsonResponse({"success": False, "errors": form.errors}, status=400)
    return JsonResponse({"success": False}, status=400)

@login_required
def delete_review(request, movie_id):
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        review = get_object_or_404(Review, user=request.user, movie_id=movie_id)
        review.delete()
        return JsonResponse({"success": True})
    return JsonResponse({"success": False}, status=400)

from django.http import JsonResponse
from .models import Review

@login_required
def update_review_from_profile(request, review_id):
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        review = get_object_or_404(Review, id=review_id, user=request.user)
        review.text = request.POST.get("text", "").strip()
        review.rating = int(request.POST.get("rating", 0))
        review.save()
        return JsonResponse({"success": True, "text": review.text, "rating": review.rating})
    return JsonResponse({"success": False}, status=400)

@login_required
def delete_review_from_profile(request, review_id):
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        review = get_object_or_404(Review, id=review_id, user=request.user)
        review.delete()
        return JsonResponse({"success": True})
    return JsonResponse({"success": False}, status=400)

def get_user_status(bonus_points):
    if bonus_points >= 1000:
        return "VIP"
    elif bonus_points >= 300:
        return "Постійний клієнт"
    else:
        return "Гість"

def get_user_status_class(bonus_points):
    if bonus_points >= 1000:
        return "vip"
    elif bonus_points >= 300:
        return "client"
    else:
        return "guest"



@login_required
def use_bonus(request):
    if request.method == "POST":
        current_balance = sum(BonusTransaction.objects.filter(user=request.user).values_list('amount', flat=True))
        if current_balance >= 50:
            # Віднімаємо бали
            new_balance = current_balance - 50
            BonusTransaction.objects.create(
                user=request.user,
                description="Використано 50 балів для знижки",
                amount=-50,
                balance_after=new_balance
            )
            # 🔹 Ставимо прапорець у сесії
            request.session["bonus_discount"] = True
        return redirect('user_dashboard')  # бо в тебе name='user_dashboard'


def add_bonus_points(user, amount, description="Нарахування бонусів"):
    last_txn = BonusTransaction.objects.filter(user=user).order_by('-date').first()
    prev_balance = last_txn.balance_after if last_txn else 0
    new_balance = prev_balance + amount

    BonusTransaction.objects.create(
        user=user,
        amount=amount,
        balance_after=new_balance,
        description=description
    )


@login_required
def edit_profile(request):
    user = request.user
    profile_updated = False
    password_changed = False

    if request.method == 'POST':
        profile_form = ProfileUpdateForm(request.POST, instance=user)
        password_form = CustomPasswordChangeForm(user, request.POST)

        if profile_form.is_valid():
            profile_form.save()
            profile_updated = True
            messages.success(request, '✅ Дані профілю оновлено.')

        if password_form.is_valid():
            password_form.save()
            update_session_auth_hash(request, password_form.user)
            password_changed = True
            messages.success(request, '🔐 Пароль успішно змінено.')

    else:
        profile_form = ProfileUpdateForm(instance=user)
        password_form = CustomPasswordChangeForm(user)

    return render(request, 'cinema/edit_profile.html', {
        'profile_form': profile_form,
        'password_form': password_form,
        'profile_updated': profile_updated,
        'password_changed': password_changed,
    })

def about_view(request):
    return render(request, 'cinema/about.html')

def contacts_view(request):
    return render(request, 'cinema/contacts.html')

from django.shortcuts import render

def faq_view(request):
    faq_data = {
        "Бронювання": [
            {"question": "Як забронювати квиток онлайн?", "answer": "Виберіть фільм, сеанс, місце, натисніть «Бронювати» та завершіть оформлення."},
            {"question": "Чи можу я змінити місце після бронювання?", "answer": "Так, змінити місце можна через касу або, якщо ще не оплачено, через особистий кабінет."},
        ],
        "Оплата": [
            {"question": "Які способи оплати підтримуються?", "answer": "Платіжна картка (Visa/MasterCard), Google Pay, Apple Pay, готівка у касі."},
            {"question": "Чи можна повернути гроші за квиток?", "answer": "Повернення можливе не пізніше ніж за 2 години до сеансу."},
        ],
        "Зали і послуги": [
            {"question": "Які є формати залів?", "answer": "У нас є зали 2D, 3D, IMAX, VIP зони та Dolby Atmos."},
            {"question": "Чи доступні фільми українською?", "answer": "Так, переважна більшість фільмів демонструються українською мовою."},
        ],
        "Інше": [
            {"question": "Чи працює кінотеатр у святкові дні?", "answer": "Так, ми працюємо щодня, включно зі святами."},
            {"question": "Чи є знижки для дітей чи студентів?", "answer": "Так, діє система знижок — деталі у касі або на сайті."},
        ]
    }
    return render(request, 'cinema/faq.html', {'faq_data': faq_data})