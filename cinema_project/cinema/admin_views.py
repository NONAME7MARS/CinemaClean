from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.conf import settings
import os
import openpyxl
from django.http import HttpResponse
from .models import Ticket, Session, Movie
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from collections import defaultdict
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
from django.db.models import Count, Sum
from django.utils.safestring import mark_safe
from datetime import date, timedelta
import json
from django.db.models import ExpressionWrapper, FloatField, F
from django.db.models.functions import ExtractHour
from django.views.decorators.csrf import csrf_exempt
import base64
from io import BytesIO
from reportlab.lib.utils import ImageReader
from django.contrib.auth.decorators import user_passes_test


def analytics_access_required(view_func):
    def check_user(user):
        return (
            user.is_superuser or
            user.groups.filter(name__in=["Директор", "Адміністратор"]).exists()
        )
    return user_passes_test(check_user)(view_func)


@analytics_access_required
def analytics_view(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    tickets = Ticket.objects.select_related('session__movie', 'session__hall')

    if start_date:
        tickets = tickets.filter(session__date__gte=start_date)
    if end_date:
        tickets = tickets.filter(session__date__lte=end_date)

    # Продажі по днях
    sales_by_date = (
        tickets
        .values("session__date")
        .annotate(total=Count("id"))
        .order_by("session__date")
    )

    # Прибуток по фільмах
    revenue_by_movie = (
        tickets
        .values("session__movie__title")
        .annotate(total=Sum("price"))
        .order_by("-total")[:5]
    )

    # Середня ціна по днях
    avg_price_by_date = (
        tickets
        .values("session__date")
        .annotate(
            avg_price=ExpressionWrapper(
                Sum("price") * 1.0 / Count("id"),
                output_field=FloatField()
            )
        )
        .order_by("session__date")
    )

    # KPI
    today = date.today()
    week_ago = today - timedelta(days=7)

    total_tickets = tickets.count()
    total_revenue = tickets.aggregate(Sum('price'))['price__sum'] or 0
    tickets_this_week = tickets.filter(session__date__gte=week_ago).count()
    most_popular_movie = (
            tickets.values(title=F('session__movie__title'))
            .annotate(count=Count('id'))
            .order_by('-count')
            .values_list('title', flat=True)
            .first() or "—"
    )

    bar_data = {
        "labels": [d["session__date"].strftime("%Y-%m-%d") for d in sales_by_date],
        "values": [d["total"] for d in sales_by_date],
    }

    pie_data = {
        "labels": [m["session__movie__title"] for m in revenue_by_movie],
        "values": [float(m["total"]) for m in revenue_by_movie],
    }

    line_chart_data = {
        "labels": [d["session__date"].strftime("%Y-%m-%d") for d in avg_price_by_date],
        "datasets": [{
            "label": "Середня ціна квитка",
            "data": [float(d["avg_price"]) for d in avg_price_by_date],
            "fill": False,
            "borderColor": "#f39c12",
            "tension": 0.3
        }]
    }
    top_movies = (
        tickets
        .values('session__movie__title')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )
    top_days = (
        tickets
        .values('session__date')
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )
    status_data = (
        tickets
        .values('status')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    status_distribution = {
        "labels": [s["status"] for s in status_data],
        "values": [s["count"] for s in status_data],
    }
    hourly_distribution = (
        tickets
        .annotate(hour=ExtractHour('session__start_time'))
        .values('hour')
        .annotate(count=Count('id'))
        .order_by('hour')
    )
    hourly_data = {
        "labels": [f"{h['hour']:02d}:00" for h in hourly_distribution],
        "values": [h['count'] for h in hourly_distribution],
    }
    hall_distribution = (
        tickets
        .values('session__hall__number')
        .annotate(count=Count('id'))
        .order_by('session__hall__number')
    )
    hall_data = {
        "labels": [f"Зал {h['session__hall__number']}" for h in hall_distribution],
        "values": [h['count'] for h in hall_distribution],
    }

    return render(request, "cinema/analytics.html", {
        "bar_data": mark_safe(json.dumps(bar_data)),
        "pie_data": mark_safe(json.dumps(pie_data)),
        "line_chart_data": mark_safe(json.dumps(line_chart_data)),
        "total_tickets": total_tickets,
        "total_revenue": total_revenue,
        "tickets_this_week": tickets_this_week,
        "most_popular_movie": most_popular_movie,
        "top_movies": top_movies,
        "top_days": top_days,
        "status_distribution": mark_safe(json.dumps(status_distribution)),
        "hall_data": mark_safe(json.dumps(hall_data)),
        "hourly_data": mark_safe(json.dumps(hourly_data)),

    })


@csrf_exempt
@analytics_access_required
def export_analytics_pdf(request):
    selected_fields = request.POST.getlist("export_fields")
    chart_data = request.POST.dict()

    chart_images = {}
    for key, val in chart_data.items():
        if key.startswith("chart_images[") and val.startswith("data:image"):
            chart_name = key.replace("chart_images[", "").replace("]", "")
            image_data = base64.b64decode(val.split(",")[1])
            chart_images[chart_name] = ImageReader(BytesIO(image_data))

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="аналітика.pdf"'

    # 📎 Завантаження шрифту з підтримкою кирилиці
    font_path = os.path.join(settings.BASE_DIR.parent, 'cinema_project', 'static', 'fonts', 'DejaVuSans.ttf')
    pdfmetrics.registerFont(TTFont("DejaVu", font_path))

    p = canvas.Canvas(response, pagesize=A4)
    p.setFont("DejaVu", 12)
    y = 800
    margin = 50

    def check_space(offset=100):
        nonlocal y
        if y < offset:
            p.showPage()
            p.setFont("DejaVu", 12)
            y = 800

    def draw_title(text):
        nonlocal y
        p.setFont("DejaVu", 13)
        p.drawString(margin, y, text)
        y -= 20
        p.setFont("DejaVu", 12)

    # 📌 KPI
    if "kpi" in selected_fields:
        check_space()
        draw_title("📌 KPI Метрики")
        total = Ticket.objects.count()
        revenue = Ticket.objects.aggregate(Sum("price"))["price__sum"] or 0
        week_ago = date.today() - timedelta(days=7)
        week_tickets = Ticket.objects.filter(session__date__gte=week_ago).count()
        popular = Ticket.objects.values(title=F("session__movie__title")).annotate(c=Count("id")).order_by("-c").first()
        popular_title = popular["title"] if popular else "—"

        p.drawString(margin + 10, y, f"Всього квитків: {total}")
        y -= 20
        p.drawString(margin + 10, y, f"Загальний дохід: {revenue:.2f} грн")
        y -= 20
        p.drawString(margin + 10, y, f"Квитків за тиждень: {week_tickets}")
        y -= 20
        p.drawString(margin + 10, y, f"Найпопулярніший фільм: {popular_title}")
        y -= 30

    # 🖼️ Вставка графіків
    chart_titles = {
        "avg_price": "📈 Середня ціна квитка по днях",
        "sales": "📆 Продажі по днях",
        "revenue": "🎬 Прибуток по фільмах",
        "status": "🧾 Розподіл за статусом",
        "halls": "🏟️ Розподіл по залах",
        "hours": "⏰ Завантаженість по годинах",
    }

    for chart_key, chart_title in chart_titles.items():
        if chart_key in selected_fields and chart_key in chart_images:
            check_space(offset=250)
            draw_title(chart_title)
            p.drawImage(chart_images[chart_key], margin, y - 200, width=400, height=200)
            y -= 220

    p.showPage()
    p.save()
    return response


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as ExcelImage
from io import BytesIO
import base64
from datetime import date, timedelta
from django.db.models import Count, Sum, F
from django.contrib.admin.views.decorators import staff_member_required
from .models import Ticket
@analytics_access_required
def export_analytics_excel(request):
    export_fields = request.POST.getlist("export_fields")
    start_date = request.POST.get("start_date")
    end_date = request.POST.get("end_date")

    tickets = Ticket.objects.select_related("session__movie", "session__hall")
    if start_date:
        tickets = tickets.filter(session__date__gte=start_date)
    if end_date:
        tickets = tickets.filter(session__date__lte=end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Аналітика"

    row = 1
    bold_font = Font(bold=True, name='Arial')  # ✅ кирилиця

    def add_title(title):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = bold_font
        row += 2

    def insert_chart(title, base64_data):
        nonlocal row
        if not base64_data:
            return
        try:
            img_data = base64.b64decode(base64_data.split(",")[1])
            img = ExcelImage(BytesIO(img_data))

            # 🔍 ЩЕ БІЛЬШИЙ РОЗМІР ДІАГРАМИ
            img.width = 800  # було 600
            img.height = 400  # було 300

            # 📌 БІЛЬШИЙ ВІДСТУП
            add_title(title)
            ws.add_image(img, f"A{row}")
            row += 22  # раніше було 22
        except Exception:
            pass

    # KPI
    if "kpi" in export_fields:
        total = tickets.count()
        revenue = tickets.aggregate(Sum("price"))["price__sum"] or 0
        week_ago = date.today() - timedelta(days=7)
        week_tickets = tickets.filter(session__date__gte=week_ago).count()
        most_popular = tickets.values("session__movie__title").annotate(c=Count("id")).order_by("-c").first()
        popular_title = most_popular["session__movie__title"] if most_popular else "—"

        add_title("📌 KPI метрики")
        ws.append(["Всього квитків", total])
        ws.append(["Загальний дохід", f"{revenue:.2f} грн"])
        ws.append(["За останній тиждень", week_tickets])
        ws.append(["Найпопулярніший фільм", popular_title])
        row = ws.max_row + 2

    # Таблиці
    if "top_movies" in export_fields:
        add_title("🎟️ Топ-10 фільмів за кількістю квитків")
        ws.append(["Фільм", "К-сть квитків"])
        for item in tickets.values("session__movie__title").annotate(c=Count("id")).order_by("-c")[:10]:
            ws.append([item["session__movie__title"], item["c"]])
        row = ws.max_row + 2

    if "top_days" in export_fields:
        add_title("📅 Топ-5 днів за продажами")
        ws.append(["Дата", "К-сть квитків"])
        for item in tickets.values("session__date").annotate(c=Count("id")).order_by("-c")[:5]:
            ws.append([item["session__date"].strftime("%Y-%m-%d"), item["c"]])
        row = ws.max_row + 2

    # Графіки
    insert_chart("📈 Середня ціна квитка", request.POST.get("chart_images[avg_price]", ""))
    insert_chart("📆 Продажі по днях", request.POST.get("chart_images[sales]", ""))
    insert_chart("🎬 Прибуток по фільмах", request.POST.get("chart_images[revenue]", ""))
    insert_chart("🧾 Розподіл за статусом", request.POST.get("chart_images[status]", ""))
    insert_chart("🏟️ Розподіл по залах", request.POST.get("chart_images[halls]", ""))
    insert_chart("⏰ Завантаженість по годинах", request.POST.get("chart_images[hours]", ""))

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="аналітика_експорт.xlsx"'
    wb.save(response)
    return response



